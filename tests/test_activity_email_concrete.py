import sys
import tempfile
import unittest
import json
import shutil
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

import auth
from database import activity_workbook, audit_log, calibration_records, concrete_records
from scripts import upload_dashboard_backup


class FakeGraphResponse:
    def __init__(self, payload=None, status=200):
        self.payload = payload or {}
        self.status = status

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def read(self):
        return json.dumps(self.payload).encode("utf-8")


class FakeAuditCollection:
    def __init__(self):
        self.inserted = None
        self.updates = []
        self.records = []

    def insert_one(self, document):
        self.inserted = dict(document)
        self.records.append(dict(document))
        return SimpleNamespace(inserted_id="mongo-id")

    def update_one(self, query, update):
        self.updates.append((query, update))

    def find(self, _query):
        records = self.records

        class Cursor(list):
            def sort(self, *_args):
                return self

        return Cursor(records)

    def update_many(self, query, update):
        self.updates.append((query, update))


class DashboardFeatureTests(unittest.TestCase):
    def test_admin_can_update_expired_calibration_in_cloud_and_local_workbooks(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            authoritative_path = temp_path / "cloud-QAQC_Master.xlsx"
            local_path = temp_path / "local-QAQC_Master.xlsx"
            workbook = Workbook()
            calibration_sheet = workbook.active
            calibration_sheet.title = "Calibration Log"
            calibration_sheet.append(
                [
                    "Calibration_ID",
                    "Equipment_Type",
                    "Calibration_Date",
                    "Next_Due_Date",
                    "Reminder_Date",
                    "Days_Until_Due",
                    "Alert_Status",
                    "Status",
                    "Certificate_No",
                    "Remarks",
                ]
            )
            calibration_sheet.append(
                [
                    "CAL-EXPIRED-001",
                    "Pressure gauge",
                    datetime.combine(date.today() - timedelta(days=730), datetime.min.time()),
                    datetime.combine(date.today() - timedelta(days=365), datetime.min.time()),
                    None,
                    -365,
                    "Overdue",
                    "Pending Calibration",
                    "OLD-CERT",
                    "Expired equipment",
                ]
            )
            workbook.create_sheet("Other Data")["A1"] = "preserved"
            workbook.save(authoritative_path)

            captured = {}

            def fake_download(destination):
                shutil.copy2(authoritative_path, destination)
                return True, "qaqc-dashboard/data/QAQC_Master.xlsx"

            def fake_upload(path, public_id):
                uploaded = load_workbook(path, data_only=True)
                captured["row"] = list(
                    uploaded["Calibration Log"].iter_rows(min_row=2, max_row=2, values_only=True)
                )[0]
                captured["other"] = uploaded["Other Data"]["A1"].value
                captured["public_id"] = public_id
                uploaded.close()
                return {"public_id": public_id}

            with patch.object(calibration_records, "LOCAL_WORKBOOK", local_path), \
                 patch.object(calibration_records, "_download_current_workbook", side_effect=fake_download), \
                 patch.object(calibration_records, "upload_master_workbook", side_effect=fake_upload):
                updated, storage = calibration_records.update_calibration_record(
                    record_id="CAL-EXPIRED-001",
                    calibration_date=date.today(),
                    next_due_date=date.today() + timedelta(days=365),
                    status="Calibrated / Active",
                    username="admin",
                    certificate_no="NEW-CERT-001",
                    notes="Calibration completed and certificate received.",
                )

            headers = [cell.value for cell in calibration_sheet[1]]
            uploaded_row = dict(zip(headers, captured["row"]))
            local_workbook = load_workbook(local_path, data_only=True)
            local_row = list(
                local_workbook["Calibration Log"].iter_rows(min_row=2, max_row=2, values_only=True)
            )[0]
            local_workbook.close()

        self.assertEqual(storage, "cloudinary")
        self.assertEqual(updated["record_id"], "CAL-EXPIRED-001")
        self.assertEqual(updated["status"], "Calibrated / Active")
        self.assertEqual(updated["days_until_due"], 365)
        self.assertEqual(updated["alert_status"], "OK")
        self.assertEqual(uploaded_row["Calibration_Date"].date(), date.today())
        self.assertEqual(uploaded_row["Next_Due_Date"].date(), date.today() + timedelta(days=365))
        self.assertEqual(uploaded_row["Reminder_Date"].date(), date.today() + timedelta(days=344))
        self.assertEqual(uploaded_row["Status"], "Calibrated / Active")
        self.assertEqual(uploaded_row["Certificate_No"], "NEW-CERT-001")
        self.assertIn("Calibration completed", uploaded_row["Remarks"])
        self.assertEqual(captured["other"], "preserved")
        self.assertEqual(captured["public_id"], "qaqc-dashboard/data/QAQC_Master.xlsx")
        self.assertEqual(local_row, captured["row"])

    def test_activity_workbook_configures_cloudinary_directly(self):
        upload_result = {
            "secure_url": "https://cloudinary.example/activity.xlsx",
            "public_id": "qaqc-dashboard/activity-logs/QAQC_Activity_Log.xlsx",
            "version": 123,
            "bytes": 456,
        }
        record = {
            "event_id": "event",
            "occurred_at": datetime(2026, 7, 21, 14, 5, tzinfo=timezone.utc),
            "action": "test",
        }
        captured = {}

        def fake_upload(path, **_kwargs):
            workbook = load_workbook(path, data_only=True)
            captured["sheets"] = workbook.sheetnames
            captured["event_id"] = workbook["2026-07-21"]["B5"].value
            return upload_result

        settings = {
            "CLOUDINARY_URL": "cloudinary://api-key:api-secret@cloud-name",
            "QAQC_ACTIVITY_WORKBOOK_PUBLIC_ID": upload_result["public_id"],
        }
        with patch.object(
            audit_log,
            "get_setting",
            side_effect=lambda name, default="": settings.get(name, default),
        ), patch.object(
            audit_log.cloudinary.uploader,
            "upload",
            side_effect=fake_upload,
        ) as upload:
            archive = audit_log._upload_activity_workbook([record])

        self.assertEqual(archive["public_id"], upload_result["public_id"])
        self.assertEqual(captured["sheets"], ["2026-07-21"])
        self.assertEqual(captured["event_id"], "event")
        self.assertEqual(upload.call_args.kwargs["type"], "authenticated")
        self.assertEqual(upload.call_args.kwargs["resource_type"], "raw")
        self.assertTrue(upload.call_args.kwargs["overwrite"])

    def test_activity_workbook_uses_one_sheet_per_date_and_deduplicates_events(self):
        records = [
            {
                "event_id": "event-1",
                "occurred_at": datetime(2026, 7, 21, 9, 0, tzinfo=timezone.utc),
                "username": "admin",
                "action": "sign_in",
                "details": {},
            },
            {
                "event_id": "event-1",
                "occurred_at": datetime(2026, 7, 21, 9, 0, tzinfo=timezone.utc),
                "username": "admin",
                "action": "sign_in",
                "details": {},
            },
            {
                "event_id": "event-2",
                "occurred_at": datetime(2026, 7, 22, 0, 1, tzinfo=timezone.utc),
                "username": "requestor",
                "action": "view_page",
                "details": {"page": "Overview"},
            },
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "QAQC_Activity_Log.xlsx"
            result = activity_workbook.build_activity_workbook(records, workbook_path)
            workbook = load_workbook(workbook_path, data_only=True)

        self.assertEqual(result["record_count"], 2)
        self.assertEqual(workbook.sheetnames, ["2026-07-21", "2026-07-22"])
        self.assertEqual(workbook["2026-07-21"].max_row, 5)
        self.assertEqual(workbook["2026-07-22"]["B5"].value, "event-2")
        self.assertEqual(workbook["2026-07-22"].freeze_panes, "A5")

    def test_admin_approval_sends_exchange_email_and_records_delivery(self):
        users = {
            "requestor": {
                "username": "requestor",
                "email": "requestor@example.com",
                "name": "QA Requestor",
                "role": "user",
                "status": "pending",
            }
        }
        exchange_settings = {
            "QAQC_EXCHANGE_TENANT_ID": "tenant-id",
            "QAQC_EXCHANGE_CLIENT_ID": "client-id",
            "QAQC_EXCHANGE_CLIENT_SECRET": "client-secret",
            "QAQC_EXCHANGE_SENDER": "fortune.kpakue@evomeclimited.com",
            "QAQC_APP_URL": "https://qaqc.example.com",
        }
        snapshots = []
        graph_requests = []

        def fake_urlopen(request, timeout):
            self.assertEqual(timeout, 20)
            graph_requests.append(request)
            if "oauth2/v2.0/token" in request.full_url:
                return FakeGraphResponse({"access_token": "graph-token"})
            return FakeGraphResponse(status=202)

        with patch.object(auth, "_load_users", return_value=users), \
             patch.object(auth, "current_user", return_value={"username": "admin", "role": "admin", "status": "approved"}), \
             patch.object(auth, "_try_save_users", side_effect=lambda value: snapshots.append(dict(value["requestor"])) or True), \
             patch.object(auth, "get_setting", side_effect=lambda name, default="": exchange_settings.get(name, default)), \
             patch.object(auth, "urlopen", side_effect=fake_urlopen), \
             patch.object(auth, "record_activity") as activity:
            ok, message = auth.approve_user("requestor", "viewer")

        self.assertTrue(ok)
        self.assertEqual(message, "Approved and email sent.")
        self.assertEqual(len(graph_requests), 2)
        self.assertIn("login.microsoftonline.com/tenant-id", graph_requests[0].full_url)
        self.assertIn(
            "graph.microsoft.com/v1.0/users/fortune.kpakue%40evomeclimited.com/sendMail",
            graph_requests[1].full_url,
        )
        mail_payload = json.loads(graph_requests[1].data.decode("utf-8"))
        self.assertEqual(
            mail_payload["message"]["toRecipients"][0]["emailAddress"]["address"],
            "requestor@example.com",
        )
        self.assertIn("access approved", mail_payload["message"]["subject"].lower())
        self.assertIn("Assigned role: Viewer", mail_payload["message"]["body"]["content"])
        self.assertIn("https://qaqc.example.com", mail_payload["message"]["body"]["content"])
        self.assertIn("Evomec Global Services QA/QC Dashboard", mail_payload["message"]["body"]["content"])
        self.assertIn("KPAKUE FORTUNE (QA)", mail_payload["message"]["body"]["content"])
        self.assertIsNotNone(snapshots[-1]["approval_email_sent_at"])
        self.assertIsNone(snapshots[-1]["approval_email_error"])
        activity.assert_called_once()

    def test_send_email_prefers_configured_gmail(self):
        settings = {
            "QAQC_GMAIL_ADDRESS": "fortunesundayske@gmail.com",
            "QAQC_GMAIL_APP_PASSWORD": "abcd efgh ijkl mnop",
        }
        smtp = MagicMock()
        smtp.__enter__.return_value = smtp

        with patch.object(auth, "get_setting", side_effect=lambda name, default="": settings.get(name, default)), \
             patch.object(auth.smtplib, "SMTP_SSL", return_value=smtp) as smtp_ssl:
            sent = auth.send_email("requestor@example.com", "Access approved", "You have access.")

        self.assertTrue(sent)
        smtp_ssl.assert_called_once_with("smtp.gmail.com", 465, timeout=20)
        smtp.login.assert_called_once_with("fortunesundayske@gmail.com", "abcdefghijklmnop")
        message = smtp.send_message.call_args.args[0]
        self.assertEqual(message["To"], "requestor@example.com")
        self.assertEqual(message["Subject"], "Access approved")

    def test_activity_is_saved_to_mongodb_and_cloudinary_without_secrets(self):
        collection = FakeAuditCollection()
        archive = {
            "url": "https://cloudinary.example/activity.xlsx",
            "public_id": "qaqc-dashboard/activity-logs/QAQC_Activity_Log.xlsx",
            "version": 123,
        }
        with patch.object(audit_log, "ensure_activity_log", return_value=collection), \
             patch.object(
                 audit_log,
                 "_upload_activity_workbook",
                 side_effect=lambda records: {
                     **archive,
                     "event_ids": [record["event_id"] for record in records],
                     "record_count": len(records),
                     "sheet_names": ["2026-07-21"],
                 },
             ):
            saved = audit_log.record_activity(
                "update_record",
                actor={"username": "tester", "name": "Test User", "role": "user"},
                details={"record": "NCR-1", "password": "must-not-be-stored"},
            )

        self.assertTrue(saved)
        self.assertEqual(collection.inserted["username"], "tester")
        self.assertNotIn("password", collection.inserted["details"])
        self.assertEqual(collection.updates[-1][1]["$set"]["cloud_archive_status"], "archived")
        self.assertEqual(
            collection.updates[-1][1]["$set"]["cloud_archive_public_id"],
            "qaqc-dashboard/activity-logs/QAQC_Activity_Log.xlsx",
        )

    def test_admin_can_append_daily_concrete_volume(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "QAQC_Master.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Concrete Tracker"
            worksheet.append(["Pour_ID", "Project", "Date", "Location", "Volume"])
            workbook.save(workbook_path)

            with patch.object(concrete_records, "LOCAL_WORKBOOK", workbook_path), \
                 patch.object(concrete_records, "get_setting", return_value=""):
                record, storage = concrete_records.append_concrete_volume(
                    entry_date=date(2026, 7, 21),
                    project="NLNG",
                    location="Tank foundation",
                    volume=42.5,
                    username="admin",
                    notes="Daily total",
                )
                second_record, _storage = concrete_records.append_concrete_volume(
                    entry_date=date(2026, 7, 22),
                    project="  nlng  ",
                    location="Pile cap",
                    volume=18.0,
                    username="admin",
                )

            saved_workbook = load_workbook(workbook_path, data_only=True)
            saved_sheet = saved_workbook["Concrete Tracker"]
            headers = [cell.value for cell in saved_sheet[1]]
            row = {headers[index]: value for index, value in enumerate(next(saved_sheet.iter_rows(min_row=2, max_row=2, values_only=True)))}
            self.assertEqual(storage, "local")
            self.assertEqual(row["Project"], "NLNG")
            self.assertEqual(row["Volume"], 42.5)
            self.assertEqual(row["Entered_By"], "admin")
            self.assertEqual(record["Location"], "Tank foundation")
            self.assertEqual(second_record["Project"], "NLNG")
            self.assertFalse(second_record["Project_Added_To_Register"])
            project_sheet = saved_workbook["Project Register"]
            registered = [cell.value for cell in project_sheet["A"][1:] if cell.value]
            self.assertEqual(registered, ["NLNG"])
            self.assertEqual(
                [saved_sheet.cell(row=row_number, column=2).value for row_number in (2, 3)],
                ["NLNG", "NLNG"],
            )

    def test_dashboard_backup_excludes_standards_and_secrets(self):
        self.assertFalse(upload_dashboard_backup.should_include(ROOT_DIR / "assets" / "standards" / "example.pdf"))
        self.assertFalse(upload_dashboard_backup.should_include(ROOT_DIR / ".env"))
        self.assertFalse(upload_dashboard_backup.should_include(ROOT_DIR / "data" / "users.json"))
        self.assertFalse(upload_dashboard_backup.should_include(ROOT_DIR / "data" / "profile_photos" / "user.png"))
        self.assertFalse(upload_dashboard_backup.should_include(ROOT_DIR / "tmp" / "report.pdf"))
        self.assertTrue(upload_dashboard_backup.should_include(ROOT_DIR / "auth.py"))


if __name__ == "__main__":
    unittest.main()
