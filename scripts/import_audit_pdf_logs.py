from pathlib import Path
from datetime import datetime

import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
WORKBOOK = BASE_DIR / "data" / "QAQC_Master.xlsx"
KPI_PDF = Path(r"C:\Users\evome\OneDrive - EVOMEC GLOBAL SERVICES LIMITED (1)\ALL\IA NEW WAREHOUSE PROJECT\AUDIT 2026\AUDIT DOCUMENT\QAQC_KPI_KRA_Register.pdf")
CTQ_PDF = Path(r"C:\Users\evome\OneDrive - EVOMEC GLOBAL SERVICES LIMITED (1)\ALL\IA NEW WAREHOUSE PROJECT\AUDIT 2026\AUDIT DOCUMENT\CTQ Log- QUERY NEW IA WAREHOUSE.pdf")
LESSONS_PDF = Path(r"C:\Users\evome\OneDrive - EVOMEC GLOBAL SERVICES LIMITED (1)\ALL\IA NEW WAREHOUSE PROJECT\AUDIT 2026\AUDIT DOCUMENT\Copy of Lessons Learnt and organizational knowledge log - New IA Warehouse Project Package A 22-08-25.pdf")


def clean(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", "\n").split())


def multiline(value):
    if value is None:
        return ""
    return "\n".join(line.strip() for line in str(value).splitlines() if line.strip())


def extract_tables(path):
    rows = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                rows.extend(table)
    return rows


def parse_date(value):
    value = clean(value)
    if not value or value.lower() == "pending":
        return ""
    for fmt in ("%d-%b-%y", "%d/%m/%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    return value


def reset_sheet(wb, name):
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)
    return ws


def write_table(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="0B1F3A")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D7DEE8")
    border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column_cells[:80])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)


def kpi_rows():
    table = extract_tables(KPI_PDF)[1:]
    rows = []
    current_kra = ""
    for kra, kpi, target, frequency in table:
        current_kra = clean(kra) or current_kra
        rows.append([current_kra, clean(kpi), clean(target), multiline(frequency), ""])
    return rows


def ctq_rows():
    table = extract_tables(CTQ_PDF)
    headers = table[1]
    raw_rows = [dict(zip(headers, row)) for row in table[2:] if clean(row[0])]
    rows = []
    for item in raw_rows:
        title = multiline(item.get("CTQ TITLE/DESCRIPTION"))
        status = clean(item.get("STATUS"))
        status_for_dashboard = status if status and status.upper() != "AFC" else "Open"
        issue_date = parse_date(item.get("ISSUE DATE"))
        response_date = parse_date(item.get("RESPONSE DATE"))
        rows.append([
            clean(item.get("CTQ No")),
            "IA Warehouse",
            clean(item.get("DISCIPLINE")).title(),
            clean(item.get("Ref No")),
            title,
            "",
            "",
            "",
            "",
            "",
            "",
            status_for_dashboard,
            issue_date,
            "",
            clean(item.get("S/No")),
            clean(item.get("Ref No")),
            clean(item.get("REV")),
            status,
            response_date,
            clean(item.get("CONTRACT NO")),
            multiline(item.get("REMARK")),
        ])
    return rows


def lesson_rows():
    tables = extract_tables(LESSONS_PDF)
    rows = []
    for row in tables[5:]:
        if not row:
            continue
        if clean(row[0]) in {"S/N", "Constr. Mgr:", "Lessons Learned workshop attendees:", "Key", "LL - Lessons Learnt", "BP - Best Practice"}:
            continue
        if not clean(row[0]) and not clean(row[1]):
            continue
        if len(row) == 10:
            lesson_id, title, focus, description, _, _, root, recommendation, owner, target = row
        elif len(row) >= 8:
            lesson_id, title, focus, description, root, recommendation, owner, target = row[:8]
        else:
            continue
        lesson_id = clean(lesson_id)
        title = clean(title)
        description = multiline(description)
        if not any([lesson_id, title, description]):
            continue
        rows.append([
            lesson_id or f"LL-{len(rows) + 1}",
            "IA Warehouse",
            clean(focus) or "Project Learning",
            title,
            description,
            multiline(root),
            multiline(recommendation),
            clean(owner),
            clean(target),
            "Imported from lessons learned PDF",
            "New IA Warehouse Project Package A",
            datetime(2025, 4, 30),
        ])
    return rows


def main():
    wb = load_workbook(WORKBOOK)

    write_table(
        reset_sheet(wb, "KPI KRA Register"),
        ["KRA", "KPI", "Target", "Frequency", "Current Performance"],
        kpi_rows(),
    )

    write_table(
        reset_sheet(wb, "CTQ Log"),
        [
            "CTQ_ID", "Project", "Discipline", "Activity", "CTQ Description",
            "Acceptance Criteria", "Specification Reference", "Inspection Method",
            "Frequency", "Target Value", "Actual Value", "Status", "Date",
            "Responsible Inspector", "Source S/No", "Source Ref No", "Revision",
            "Source Status", "Response Date", "Contract No", "Remark",
        ],
        ctq_rows(),
    )

    write_table(
        reset_sheet(wb, "Lessons Learned"),
        [
            "Lesson_ID", "Project", "Category", "Lesson", "Description",
            "Root_Cause", "Recommendation", "Action_Owner", "Target_Date",
            "Impact", "Source_Project", "Date_Logged",
        ],
        lesson_rows(),
    )

    wb.save(WORKBOOK)


if __name__ == "__main__":
    main()
