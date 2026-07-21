"""Persistent manual concrete-volume entry support."""

from copy import copy
from datetime import datetime, timezone
from pathlib import Path
import re
import shutil
import tempfile
import unicodedata
from urllib.request import urlopen
import uuid

from openpyxl import load_workbook

from database.cloudinary_storage import (
    DEFAULT_MASTER_WORKBOOK_PUBLIC_ID,
    get_master_workbook_reference,
    upload_master_workbook,
)
from database.settings import get_setting


BASE_DIR = Path(__file__).resolve().parents[1]
LOCAL_WORKBOOK = BASE_DIR / "data" / "QAQC_Master.xlsx"
SHEET_NAME = "Concrete Tracker"
PROJECT_REGISTER_SHEET = "Project Register"
REQUIRED_COLUMNS = [
    "Pour_ID", "Project", "Date", "Location", "Volume",
    "Entered_By", "Entered_At", "Entry_Notes",
]


def _project_key(value):
    text = unicodedata.normalize("NFKC", str(value or ""))
    return re.sub(r"\s+", " ", text).strip().casefold()


def _clean_project_name(value):
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", str(value or ""))).strip()


def _header_map(worksheet):
    return {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def _canonical_projects(workbook):
    """Return normalized-to-display project names, preferring Project Register labels."""
    projects = {}
    preferred_sheets = [PROJECT_REGISTER_SHEET, SHEET_NAME]
    remaining_sheets = [name for name in workbook.sheetnames if name not in preferred_sheets]
    for sheet_name in preferred_sheets + remaining_sheets:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        headers = _header_map(worksheet)
        project_column = headers.get("Project")
        if not project_column:
            continue
        for row in range(2, worksheet.max_row + 1):
            display_name = _clean_project_name(worksheet.cell(row=row, column=project_column).value)
            key = _project_key(display_name)
            if key and key not in projects:
                projects[key] = display_name
    return projects


def _copy_row_style(worksheet, source_row, target_row, max_column):
    if source_row < 1:
        return
    for column in range(1, max_column + 1):
        source = worksheet.cell(row=source_row, column=column)
        target = worksheet.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)


def _ensure_project_registered(workbook, project):
    worksheet = (
        workbook[PROJECT_REGISTER_SHEET]
        if PROJECT_REGISTER_SHEET in workbook.sheetnames
        else workbook.create_sheet(PROJECT_REGISTER_SHEET)
    )
    headers = _header_map(worksheet)
    if "Project" not in headers:
        next_column = max(headers.values(), default=0) + 1
        worksheet.cell(row=1, column=next_column, value="Project")
        headers["Project"] = next_column
    project_column = headers["Project"]
    project_key = _project_key(project)
    for row in range(2, worksheet.max_row + 1):
        if _project_key(worksheet.cell(row=row, column=project_column).value) == project_key:
            return False

    row_number = worksheet.max_row + 1
    _copy_row_style(worksheet, row_number - 1, row_number, max(worksheet.max_column, project_column))
    worksheet.cell(row=row_number, column=project_column, value=project)
    return True


def _download_current_workbook(destination):
    cloudinary_configured = bool(str(get_setting("CLOUDINARY_URL", "")).strip())
    public_id = str(
        get_setting("QAQC_MASTER_WORKBOOK_PUBLIC_ID", DEFAULT_MASTER_WORKBOOK_PUBLIC_ID)
    ).strip() or DEFAULT_MASTER_WORKBOOK_PUBLIC_ID
    if cloudinary_configured:
        try:
            reference = get_master_workbook_reference(public_id)
            with urlopen(reference["url"], timeout=30) as response, destination.open("wb") as output:
                shutil.copyfileobj(response, output)
            return True, public_id
        except Exception:
            pass
    if not LOCAL_WORKBOOK.exists():
        raise RuntimeError("The QA/QC master workbook is unavailable.")
    shutil.copy2(LOCAL_WORKBOOK, destination)
    return cloudinary_configured, public_id


def list_concrete_projects():
    """Read unique canonical project names from the current authoritative workbook."""
    with tempfile.TemporaryDirectory(prefix="qaqc-projects-") as temp_dir:
        workbook_path = Path(temp_dir) / "QAQC_Master.xlsx"
        _download_current_workbook(workbook_path)
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            projects = _canonical_projects(workbook)
        finally:
            workbook.close()
    return sorted(projects.values(), key=str.casefold)


def list_concrete_projects_from_data(data):
    """Return canonical project names from workbook frames already loaded by a page."""
    if not isinstance(data, dict):
        return []

    projects = {}
    preferred_sheets = [PROJECT_REGISTER_SHEET, SHEET_NAME]
    remaining_sheets = [name for name in data if name not in preferred_sheets]
    for sheet_name in preferred_sheets + remaining_sheets:
        frame = data.get(sheet_name)
        if frame is None or not hasattr(frame, "columns") or "Project" not in frame.columns:
            continue
        for value in frame["Project"].dropna().tolist():
            display_name = _clean_project_name(value)
            key = _project_key(display_name)
            if key and key not in projects:
                projects[key] = display_name
    return sorted(projects.values(), key=str.casefold)


def append_concrete_volume(*, entry_date, project, location, volume, username, notes=""):
    project = _clean_project_name(project)
    location = str(location or "").strip()
    username = str(username or "admin").strip()
    notes = str(notes or "").strip()
    volume = float(volume)
    if not project:
        raise ValueError("Project is required.")
    if not location:
        raise ValueError("Location or work area is required.")
    if volume <= 0:
        raise ValueError("Concrete volume must be greater than zero.")

    entered_at = datetime.now(timezone.utc)
    pour_id = f"P-MAN-{entry_date:%Y%m%d}-{uuid.uuid4().hex[:6].upper()}"
    record = {
        "Pour_ID": pour_id,
        "Project": project,
        "Date": datetime.combine(entry_date, datetime.min.time()),
        "Location": location,
        "Volume": volume,
        "Entered_By": username,
        "Entered_At": entered_at.replace(tzinfo=None),
        "Entry_Notes": notes,
    }

    with tempfile.TemporaryDirectory(prefix="qaqc-concrete-") as temp_dir:
        workbook_path = Path(temp_dir) / "QAQC_Master.xlsx"
        upload_to_cloud, public_id = _download_current_workbook(workbook_path)
        workbook = load_workbook(workbook_path)
        worksheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.create_sheet(SHEET_NAME)

        canonical_projects = _canonical_projects(workbook)
        project = canonical_projects.get(_project_key(project), project)
        project_registered = _ensure_project_registered(workbook, project)
        record["Project"] = project

        headers = _header_map(worksheet)
        for column in REQUIRED_COLUMNS:
            if column not in headers:
                next_column = max(headers.values(), default=0) + 1
                worksheet.cell(row=1, column=next_column, value=column)
                headers[column] = next_column

        row_number = worksheet.max_row + 1
        _copy_row_style(worksheet, row_number - 1, row_number, max(headers.values()))
        for column, value in record.items():
            worksheet.cell(row=row_number, column=headers[column], value=value)
        worksheet.cell(row=row_number, column=headers["Date"]).number_format = "yyyy-mm-dd"
        worksheet.cell(row=row_number, column=headers["Entered_At"]).number_format = "yyyy-mm-dd hh:mm:ss"
        workbook.save(workbook_path)

        storage = "local"
        if upload_to_cloud:
            upload_master_workbook(workbook_path, public_id=public_id)
            storage = "cloudinary"
        try:
            shutil.copy2(workbook_path, LOCAL_WORKBOOK)
        except OSError:
            # Hosted deployments can be read-only; the Cloudinary version is authoritative.
            pass

    record["Date"] = record["Date"].date().isoformat()
    record["Entered_At"] = entered_at.isoformat()
    record["Project_Added_To_Register"] = project_registered
    return record, storage
