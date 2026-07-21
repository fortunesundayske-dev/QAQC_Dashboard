"""Authoritative calibration-record updates for the QA/QC master workbook."""

from __future__ import annotations

from datetime import date, datetime, time, timedelta
from pathlib import Path
import shutil
import tempfile
from urllib.request import urlopen

from openpyxl import load_workbook

from database.cloudinary_storage import (
    DEFAULT_MASTER_WORKBOOK_PUBLIC_ID,
    get_master_workbook_reference,
    upload_master_workbook,
)
from database.settings import get_setting


BASE_DIR = Path(__file__).resolve().parents[1]
LOCAL_WORKBOOK = BASE_DIR / "data" / "QAQC_Master.xlsx"
SHEET_NAME = "Calibration Log"


def _header_map(worksheet):
    return {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def _as_date(value, field_name):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(str(value)).date()
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} must be a valid date.") from exc


def _download_current_workbook(destination):
    cloudinary_configured = bool(str(get_setting("CLOUDINARY_URL", "")).strip())
    public_id = str(
        get_setting("QAQC_MASTER_WORKBOOK_PUBLIC_ID", DEFAULT_MASTER_WORKBOOK_PUBLIC_ID)
    ).strip() or DEFAULT_MASTER_WORKBOOK_PUBLIC_ID

    if cloudinary_configured:
        try:
            reference = get_master_workbook_reference(public_id)
            with urlopen(reference["url"], timeout=30) as response, destination.open("wb") as output:
                shutil.copyfileobj(response, output)
        except Exception as exc:
            raise RuntimeError(
                "The latest Cloudinary master workbook could not be downloaded; no calibration change was saved."
            ) from exc
        return True, public_id

    if not LOCAL_WORKBOOK.exists():
        raise RuntimeError("The QA/QC master workbook is unavailable.")
    shutil.copy2(LOCAL_WORKBOOK, destination)
    return False, public_id


def _alert_status(next_due_date):
    days_until_due = (next_due_date - date.today()).days
    if days_until_due < 0:
        return days_until_due, "Overdue"
    if days_until_due == 21:
        return days_until_due, "Due in 21 days"
    if days_until_due < 21:
        return days_until_due, "Due soon"
    return days_until_due, "OK"


def update_calibration_record(
    *,
    record_id,
    calibration_date,
    next_due_date,
    status,
    username,
    certificate_no="",
    notes="",
):
    """Update one calibration record in Cloudinary and the local workbook copy."""
    record_id = str(record_id or "").strip()
    status = str(status or "").strip()
    username = str(username or "admin").strip()
    certificate_no = str(certificate_no or "").strip()
    notes = str(notes or "").strip()
    calibrated_on = _as_date(calibration_date, "Calibration date")
    due_on = _as_date(next_due_date, "Next due date")

    if not record_id:
        raise ValueError("An equipment record is required.")
    if not status:
        raise ValueError("Equipment status is required.")
    if due_on <= calibrated_on:
        raise ValueError("Next due date must be after the calibration date.")

    with tempfile.TemporaryDirectory(prefix="qaqc-calibration-") as temp_dir:
        workbook_path = Path(temp_dir) / "QAQC_Master.xlsx"
        upload_to_cloud, public_id = _download_current_workbook(workbook_path)
        workbook = load_workbook(workbook_path)
        if SHEET_NAME not in workbook.sheetnames:
            workbook.close()
            raise RuntimeError("The Calibration Log sheet is missing from the master workbook.")

        worksheet = workbook[SHEET_NAME]
        headers = _header_map(worksheet)
        if "Calibration_ID" not in headers:
            workbook.close()
            raise RuntimeError("The Calibration Log sheet has no Calibration_ID column.")

        matching_rows = [
            row_number
            for row_number in range(2, worksheet.max_row + 1)
            if str(worksheet.cell(row=row_number, column=headers["Calibration_ID"]).value or "").strip()
            == record_id
        ]
        if not matching_rows:
            workbook.close()
            raise ValueError(f"Calibration record {record_id} was not found.")
        if len(matching_rows) > 1:
            workbook.close()
            raise RuntimeError(f"Calibration record {record_id} is duplicated in the master workbook.")

        required_columns = [
            "Calibration_Date",
            "Next_Due_Date",
            "Reminder_Date",
            "Days_Until_Due",
            "Alert_Status",
            "Status",
            "Certificate_No",
            "Remarks",
        ]
        for column in required_columns:
            if column not in headers:
                next_column = max(headers.values(), default=0) + 1
                worksheet.cell(row=1, column=next_column, value=column)
                headers[column] = next_column

        row_number = matching_rows[0]
        previous = {
            "status": worksheet.cell(row=row_number, column=headers["Status"]).value,
            "calibration_date": worksheet.cell(row=row_number, column=headers["Calibration_Date"]).value,
            "next_due_date": worksheet.cell(row=row_number, column=headers["Next_Due_Date"]).value,
            "certificate_no": worksheet.cell(row=row_number, column=headers["Certificate_No"]).value,
        }
        days_until_due, alert_status = _alert_status(due_on)
        date_values = {
            "Calibration_Date": calibrated_on,
            "Next_Due_Date": due_on,
            "Reminder_Date": due_on - timedelta(days=21),
        }
        for column, value in date_values.items():
            cell = worksheet.cell(row=row_number, column=headers[column], value=datetime.combine(value, time.min))
            cell.number_format = "yyyy-mm-dd"
        worksheet.cell(row=row_number, column=headers["Days_Until_Due"], value=days_until_due)
        worksheet.cell(row=row_number, column=headers["Alert_Status"], value=alert_status)
        worksheet.cell(row=row_number, column=headers["Status"], value=status)
        worksheet.cell(row=row_number, column=headers["Certificate_No"], value=certificate_no)

        if notes:
            remarks_cell = worksheet.cell(row=row_number, column=headers["Remarks"])
            existing_remarks = str(remarks_cell.value or "").strip()
            update_note = f"[{date.today().isoformat()} by {username}] {notes}"
            remarks_cell.value = f"{existing_remarks}\n{update_note}".strip()

        workbook.save(workbook_path)
        workbook.close()

        storage = "local"
        if upload_to_cloud:
            upload_master_workbook(workbook_path, public_id=public_id)
            storage = "cloudinary"
        try:
            shutil.copy2(workbook_path, LOCAL_WORKBOOK)
        except OSError:
            # Hosted deployments can be read-only; Cloudinary remains authoritative.
            pass

    return {
        "record_id": record_id,
        "status": status,
        "calibration_date": calibrated_on.isoformat(),
        "next_due_date": due_on.isoformat(),
        "reminder_date": (due_on - timedelta(days=21)).isoformat(),
        "days_until_due": days_until_due,
        "alert_status": alert_status,
        "certificate_no": certificate_no,
        "previous": {
            key: value.isoformat() if isinstance(value, (date, datetime)) else str(value or "")
            for key, value in previous.items()
        },
    }, storage
