"""Excel workbook generation for the persistent QA/QC activity archive."""

from collections import defaultdict
from copy import copy
from datetime import datetime, timezone
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ACTIVITY_COLUMNS = [
    ("Occurred At (UTC)", "occurred_at"),
    ("Event ID", "event_id"),
    ("Username", "username"),
    ("Name", "name"),
    ("Email", "email"),
    ("Role", "role"),
    ("Action", "action"),
    ("Category", "category"),
    ("Page", "page"),
    ("Target", "target"),
    ("Status", "status"),
    ("Details", "details"),
]

TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBTITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN_BLUE = Side(style="thin", color="A9C4D9")


def _as_utc(value):
    if not isinstance(value, datetime):
        raise ValueError("Every activity record must have a valid occurred_at timestamp.")
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _detail_text(value):
    if not value:
        return ""
    if isinstance(value, dict):
        return json.dumps(value, default=str, sort_keys=True, separators=(", ", ": "))
    return str(value)


def _deduplicated_records(records):
    """Return one stable record per event ID, ordered by UTC timestamp."""
    by_event = {}
    for index, source in enumerate(records):
        record = dict(source)
        record.pop("_id", None)
        occurred_at = _as_utc(record.get("occurred_at"))
        event_id = str(record.get("event_id") or f"legacy-{index}-{occurred_at.timestamp()}")
        record["event_id"] = event_id
        record["occurred_at"] = occurred_at
        by_event[event_id] = record
    return sorted(by_event.values(), key=lambda item: (item["occurred_at"], item["event_id"]))


def build_activity_workbook(records, destination):
    """Build one formatted workbook with a worksheet for each UTC activity date."""
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    records = _deduplicated_records(records)
    grouped = defaultdict(list)
    for record in records:
        grouped[record["occurred_at"].date().isoformat()].append(record)

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "QA/QC Dashboard Activity Log"
    workbook.properties.subject = "Daily user activity audit trail"
    workbook.properties.creator = "Evomec QA/QC Dashboard"
    workbook.properties.description = (
        "One UTC-dated worksheet per day. Event ID is the unique activity key."
    )

    if not grouped:
        grouped[datetime.now(timezone.utc).date().isoformat()] = []

    for day, day_records in grouped.items():
        worksheet = workbook.create_sheet(day)
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A5"

        worksheet.merge_cells("A1:L1")
        worksheet["A1"] = f"QA/QC Daily Activity Log — {day} (UTC)"
        worksheet["A1"].fill = TITLE_FILL
        worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
        worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 28

        worksheet.merge_cells("A2:L2")
        worksheet["A2"] = (
            f"{len(day_records)} recorded activit{'y' if len(day_records) == 1 else 'ies'} · "
            "Event IDs prevent duplicate rows"
        )
        worksheet["A2"].fill = SUBTITLE_FILL
        worksheet["A2"].font = Font(color="17365D", italic=True)
        worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[2].height = 20

        headers = [label for label, _key in ACTIVITY_COLUMNS]
        for column, label in enumerate(headers, start=1):
            cell = worksheet.cell(row=4, column=column, value=label)
            cell.fill = HEADER_FILL
            cell.font = copy(WHITE_FONT)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=THIN_BLUE)
        worksheet.row_dimensions[4].height = 30

        for row_index, record in enumerate(day_records, start=5):
            values = []
            for _label, key in ACTIVITY_COLUMNS:
                value = record.get(key, "")
                if key == "occurred_at":
                    value = record["occurred_at"].replace(tzinfo=None)
                elif key == "details":
                    value = _detail_text(value)
                elif value is None:
                    value = ""
                elif not isinstance(value, (str, int, float, bool, datetime)):
                    value = str(value)
                values.append(value)
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row_index, column=column, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=column == 12)
                cell.border = Border(bottom=Side(style="hair", color="D9E2F3"))
            worksheet.cell(row=row_index, column=1).number_format = "yyyy-mm-dd hh:mm:ss.000"

        if day_records:
            table = Table(displayName=f"Activity_{day.replace('-', '')}", ref=f"A4:L{4 + len(day_records)}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)
            worksheet.auto_filter.ref = f"A4:L{4 + len(day_records)}"

        widths = [24, 34, 18, 24, 30, 14, 28, 20, 24, 26, 14, 60]
        for column, width in enumerate(widths, start=1):
            worksheet.column_dimensions[worksheet.cell(row=4, column=column).column_letter].width = width
        worksheet.auto_filter.ref = f"A4:L{max(4, 4 + len(day_records))}"

    workbook.save(destination)
    return {
        "path": destination,
        "event_ids": [record["event_id"] for record in records],
        "record_count": len(records),
        "sheet_names": list(grouped),
    }
